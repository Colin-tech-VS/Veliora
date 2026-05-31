"""Import CSV / Excel — acheteurs et locataires."""

from __future__ import annotations

import csv
import io
import re
from typing import Any

from crm.mandates.storage import CLIENT_SEGMENTS, create_property_client

# En-têtes reconnues (normalisées sans accents ni espaces)
FIELD_ALIASES: dict[str, list[str]] = {
    "segment": ["segment", "type", "profil", "categorie", "role"],
    "first_name": ["prenom", "firstname", "first", "prenomclient"],
    "last_name": ["nom", "lastname", "last", "nomclient", "name"],
    "phone": ["telephone", "tel", "mobile", "gsm", "phone"],
    "email": ["email", "mail", "courriel"],
    "budget_min": ["budgetmin", "budget_min", "budgetminimum", "minbudget"],
    "budget_max": ["budgetmax", "budget_max", "budgetmaximum", "maxbudget", "budget"],
    "property_type": ["typebien", "propertytype", "bien", "typologie"],
    "rooms_min": ["piecesmin", "roomsmin", "pieces", "nbpieces", "rooms"],
    "surface_min": ["surfacemin", "surface", "surfacem2", "m2"],
    "cities": ["villes", "cities", "ville", "secteur", "zones"],
    "status": ["statut", "status", "etat"],
    "notes": ["notes", "commentaire", "commentaires", "remarques"],
}


def _norm_header(value: str) -> str:
    s = (value or "").strip().lower()
    s = s.replace("é", "e").replace("è", "e").replace("ê", "e")
    s = s.replace("à", "a").replace("ù", "u").replace("ô", "o")
    s = s.replace("ï", "i").replace("ç", "c")
    return re.sub(r"[^a-z0-9]", "", s)


def _map_headers(raw_headers: list[str]) -> dict[int, str]:
    """index colonne -> clé champ interne."""
    mapping: dict[int, str] = {}
    for idx, raw in enumerate(raw_headers):
        key = _norm_header(str(raw or ""))
        if not key:
            continue
        for field, aliases in FIELD_ALIASES.items():
            if key in aliases or key == field:
                mapping[idx] = field
                break
    return mapping


def _parse_int(val: Any) -> int | None:
    if val is None or val == "":
        return None
    try:
        return int(float(str(val).replace(" ", "").replace(",", ".")))
    except (TypeError, ValueError):
        return None


def _parse_float(val: Any) -> float | None:
    if val is None or val == "":
        return None
    try:
        return float(str(val).replace(" ", "").replace(",", "."))
    except (TypeError, ValueError):
        return None


def _normalize_segment(raw: str | None, default: str | None = None) -> str:
    s = (raw or default or "acheteur").strip().lower()
    if s in CLIENT_SEGMENTS:
        return s
    if s in ("buyer", "achat", "vente", "acquereur", "acquéreur"):
        return "acheteur"
    if s in ("tenant", "location", "louer", "bailleur", "loc"):
        return "locataire"
    return default or "acheteur"


def _row_dict_from_values(values: list, col_map: dict[int, str]) -> dict:
    row: dict[str, Any] = {}
    for idx, field in col_map.items():
        if idx < len(values):
            row[field] = values[idx]
    return row


def _client_payload_from_row(row: dict, default_segment: str | None) -> dict:
    cities_raw = row.get("cities")
    if isinstance(cities_raw, list):
        cities = [str(c).strip() for c in cities_raw if str(c).strip()]
    else:
        cities = [
            c.strip()
            for c in str(cities_raw or "").replace(";", ",").split(",")
            if c.strip()
        ]
    segment = _normalize_segment(
        str(row.get("segment") or "") if row.get("segment") is not None else None,
        default_segment,
    )
    return {
        "segment": segment,
        "first_name": str(row.get("first_name") or "").strip(),
        "last_name": str(row.get("last_name") or "").strip(),
        "phone": str(row.get("phone") or "").strip(),
        "email": str(row.get("email") or "").strip(),
        "budget_min": _parse_int(row.get("budget_min")),
        "budget_max": _parse_int(row.get("budget_max")),
        "property_type": str(row.get("property_type") or "").strip(),
        "rooms_min": _parse_int(row.get("rooms_min")),
        "surface_min": _parse_float(row.get("surface_min")),
        "cities": cities,
        "status": str(row.get("status") or "actif").strip() or "actif",
        "notes": str(row.get("notes") or "").strip(),
    }


def parse_csv_bytes(data: bytes) -> list[dict]:
    text = data.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    rows_iter = iter(reader)
    try:
        headers = next(rows_iter)
    except StopIteration:
        return []
    col_map = _map_headers(headers)
    if not col_map:
        raise ValueError(
            "En-têtes non reconnues. Utilisez au minimum : segment, prenom, nom, email"
        )
    out: list[dict] = []
    for values in rows_iter:
        if not any(str(v).strip() for v in values):
            continue
        out.append(_row_dict_from_values(values, col_map))
    return out


def parse_xlsx_bytes(data: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError(
            "Import Excel indisponible — installez openpyxl : pip install openpyxl"
        ) from exc
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(c or "") for c in rows[0]]
    col_map = _map_headers(headers)
    if not col_map:
        raise ValueError(
            "En-têtes non reconnues. Première ligne : segment, prenom, nom, email…"
        )
    out: list[dict] = []
    for values in rows[1:]:
        if not values or not any(v is not None and str(v).strip() for v in values):
            continue
        vals = ["" if v is None else v for v in values]
        out.append(_row_dict_from_values(vals, col_map))
    return out


def import_clients_from_rows(
    agency_id: str,
    rows: list[dict],
    *,
    default_segment: str | None = None,
) -> dict:
    created = 0
    skipped = 0
    errors: list[dict] = []
    for line_no, row in enumerate(rows, start=2):
        try:
            payload = _client_payload_from_row(row, default_segment)
            if not (
                payload["first_name"]
                or payload["last_name"]
                or payload["email"]
                or payload["phone"]
            ):
                skipped += 1
                continue
            create_property_client(agency_id, payload)
            created += 1
        except ValueError as exc:
            errors.append({"line": line_no, "error": str(exc)})
        except Exception as exc:
            errors.append({"line": line_no, "error": str(exc)})
    return {
        "created": created,
        "skipped": skipped,
        "errors": errors,
        "total_rows": len(rows),
    }


IMPORT_TEMPLATE_CSV = (
    "segment;prenom;nom;email;telephone;budget_min;budget_max;"
    "type_bien;pieces_min;surface_min;villes;notes\r\n"
    "acheteur;Marie;Dupont;marie@exemple.fr;0612345678;200000;350000;"
    "Appartement;3;65;Lyon,Villeurbanne;Recherche T3\r\n"
    "locataire;Paul;Martin;paul@exemple.fr;0698765432;800;1200;"
    "Appartement;2;45;Lyon;Disponible mars\r\n"
)
